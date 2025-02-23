import os
import sys
import xlrd
import xlwt
import datetime
import requests
import pandas as pd

from typing import Dict, List, Optional
from openpyxl import load_workbook
from sqlalchemy import create_engine


# retrieve the RUNENV
RUNENV = os.getenv("RUNENV", "local")

# force stdout flushing
sys.stdout.reconfigure(line_buffering=True)

if RUNENV == "local-docker" or RUNENV == "prod":
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    print("PATH PRINT: ", os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils import PostgresConnector


def create_output_folder() -> None:
    """
    Creates an 'output' directory in the current working directory if it doesn't already exist.
    """
    output_directory = os.getcwd() + '/' + 'output'
    if not os.path.isdir(output_directory):
        os.mkdir(output_directory)


def download_workout_sheet(workout_sheet_url: str, workout_sheet_xlsx: str) -> None:
    """
    Downloads the workout sheet from a given URL and saves it as an XLSX file.

    Args:
        workout_sheet_url (str): The URL of the workout sheet to be downloaded.
        workout_sheet_xlsx (str): The file path to save the downloaded XLSX file.
    """
    print("Downloading workout sheet...")
    response = requests.get(workout_sheet_url, stream=True)
    if response.status_code == 200:
        with open(workout_sheet_xlsx, "wb") as file:
            for chunk in response.iter_content(chunk_size=1024):
                file.write(chunk)
        print(f"Download complete: {workout_sheet_xlsx}")
    else:
        print(f"Failed to download file. Status code: {response.status_code}")


def convert_xlsx_to_xls(xlsx_file: str, xls_file: str) -> None:
    """
    Converts an XLSX file to an XLS file format.

    Args:
        xlsx_file (str): The path of the XLSX file to be converted.
        xls_file (str): The path to save the converted XLS file.
    """
    # load the .xlsx file
    print(f"Converting workout sheet from .xlsx to xls...")
    workbook = load_workbook(xlsx_file)
    
    # create a new .xls file
    new_workbook = xlwt.Workbook()

    # loop through the sheets in the .xlsx file
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        new_sheet = new_workbook.add_sheet(sheet_name)

        # copy each row and cell from the .xlsx sheet to the .xls sheet
        for row_index, row in enumerate(sheet.iter_rows(), start=0):
            for col_index, cell in enumerate(row, start=0):
                new_sheet.write(row_index, col_index, cell.value)
    
    # save the new .xls file
    new_workbook.save(xls_file)
    print("Workout sheet converted successfully!")


def extract_weight_kg(value: str) -> tuple[Optional[float], Optional[str]]:
    """
    Extracts the weight in kg from a given string value.

    Args:
        value (str): The string containing the weight information.

    Returns:
        tuple: A tuple containing the weight as a float and the unit as a string ('kg').
    """
    if "Bodyweight" in value:
        return 0.00, "kg"  # assign 0.00 kg for bodyweight
    elif "kg" in value:
        return float(value.split("kg")[0]), "kg"  # extract numeric part before 'kg'
    return None, None  # fallback in case of unexpected values


def read_workbook(workout_sheet: str) -> Optional[Dict[str, List[List]]]:
    """
    Reads the workout sheet and processes it into a structured dictionary format.

    Args:
        workout_sheet (str): The path to the .xls workout sheet.

    Returns:
        Optional[Dict[str, List[List]]]: A dictionary where keys are years and values are lists of workout data for each year.
    """
    print(f"Loading workbook {workout_sheet}...")
    wb = xlrd.open_workbook(workout_sheet)
    print(f"Workbook '{workout_sheet}' loaded successfully!")
    available_years = wb.sheet_names()
    output_files = []
    final_dictionary = {}
    for year in available_years:
        # skipping 2023 as there is no data for this year
        if year == '2023':
            continue
        
        # read the sheet for each year
        current_sheet = wb.sheet_by_name(year)
        
        # creating the lists of acceptable values
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        # these cells are skipabble as they are the main headers of each session
        cells_to_skip = ['Exercise', 'Weight', 'Grip', 'Amount', 'Rest (s)']
        # the workout sections are entered manually as there will be added complexity to what I train
        workout_sections = ['Warm-Up', 'Pull', 'Push', 'Legs', 'Weights', 'Floor Routine', 'Bar Routine', 'Pull/Push Pyramid', 'Pull/Push']
        current_month = ''
        current_week_day = ''
        current_day = ''
        current_workout_type = ''
        current_workout_section = ''
        workout_counter = 0
        weight = 'Bodyweight' # default value
        counter = 0
        set_counter = 0
        temp_list = []
        final_list = []
        deload = 0 # default value

        # the output file will be a .txt containing a list of lists
        output_file_name = f'./output/workout_journal_{year}.txt'

        '''
        The excel spreadsheet containing the workout journal always follows the same structure, it will not change
        This way we can hardcode some relations between certain values
        Date -> Deload -> Weekday -> Workout Type -> Workout Counter
        Workout Section
        -> Exercise -> Weight -> Grip -> Amount -> Rest (s)
        '''

        with open(output_file_name, 'w') as f:
            for i in range(1, current_sheet.nrows):
                for j in range(current_sheet.ncols):
                    current_cell = current_sheet.cell_value(i, j)
                    if(current_cell in months):
                        current_month = current_cell
                        # print(f"CURRENT_MONTH: {current_month}")
                    if(current_cell in weekdays):
                        current_week_day = current_cell
                        # print(f"CURRENT_WEEK_DAY: {current_week_day}")
                        if (current_sheet.cell_value(i, j-1) == 'Deload'):
                            deload = 1
                        else:
                            deload = 0
                        current_day = str(datetime.datetime(*xlrd.xldate_as_tuple(current_sheet.cell_value(i, j-2), wb.datemode)))[:10]
                        current_day = current_day[8:10] + current_day[4] + current_day[5:7] + current_day[4] + current_day[0:4]
                        # print(f"CURRENT_DAY: {current_day}")
                        current_workout_type = current_sheet.cell_value(i, j+1)
                        # print(f"CURRENT_WORKOUT_TYPE: {current_workout_type}")
                    if(isinstance(current_cell, str)):
                        if('/365' in current_cell):
                            workout_counter = current_cell
                            # print(f"WORKOUT_COUNTER: {workout_counter}")
                        if('kg' in current_cell):
                            weight = current_cell
                        else:
                            weight = 'Bodyweight'
                        if(current_sheet.cell_value(i, j-1) == weight):
                            grip = current_cell
                    if (current_cell != current_month and current_cell != current_week_day and current_cell != current_day and current_cell != workout_counter and current_cell != current_workout_type and current_cell != 'Deload'):
                        # check for length 7 as we want to skip the date cell. We already obtain the value above.                
                        if(len(str(current_cell)) == 7  and current_sheet.cell_value(i, j+2) in weekdays):
                            continue
                        if(current_cell in cells_to_skip):
                            continue
                        if(isinstance(current_cell, str)):
                            if 'Total' in current_cell:
                                break # we want to skip this row, since I don't want aggregated metrics in the same table as the raw data
                        if(current_cell in workout_sections):
                            current_workout_section = current_cell
                            set_counter = 0
                            # print(f"CURRENT SECTION: {current_workout_section}")
                            continue
                        if(type(current_cell) == float):
                            current_cell = int(current_cell)
                        # creating an array of arrays with the other cells, append current cell
                        if(current_cell != current_workout_section):
                            if(current_cell == ''):
                                continue
                            else:
                                counter = counter + 1
                                # print(counter, current_cell)
                                if(counter <= 5):
                                    temp_list.append(current_cell)
                                if(counter == 5):
                                    counter = 0
                                    set_counter = set_counter + 1
                                    final_list.append([current_month, current_week_day, current_day, deload, workout_counter, set_counter, current_workout_type, current_workout_section] + temp_list)
                                    # print(f"FINAL LIST: {final_list}")
                                    temp_list = []
            f.write(str(final_list))
            final_dictionary[year] = final_list
            output_files.append(output_file_name)

    return final_dictionary if final_dictionary else None


def insert_dataframes_into_postgres(processed_dataframes: Dict[str, pd.DataFrame], schema: str, table_name: str) -> None:
    """
    Inserts the processed workout data into a PostgreSQL database.

    Args:
        processed_dataframes (Dict[str, pd.DataFrame]): A dictionary of dataframes to be inserted into PostgreSQL.
        schema (str): The database schema where the table resides.
        table_name (str): The name of the table in the PostgreSQL database.
    """
    if RUNENV == "local-docker" or RUNENV == "prod":
        script_dir = os.path.dirname(os.path.abspath(__file__))
        sql_file_path = os.path.join(script_dir, "sql", "CREATE_workout_journal.sql")
    else:
        sql_file_path = "sql/CREATE_workout_journal.sql"
    with PostgresConnector() as pg_conn:
        drop_table_query = f"DROP TABLE IF EXISTS {schema}.{table_name};"
        create_table_query = open(sql_file_path).read().format(schema=schema,table_name=table_name)
        print(f"Deleting table {schema}.{table_name}...")
        pg_conn.execute_query(drop_table_query)
        print("Table succesfully deleted!")
        print(f"Creating table {schema}.{table_name}...")
        pg_conn.execute_query(create_table_query)
        print("Table created successfully!")
        print("Creating sqlalchemy engine with the existing psycopg connection...")
        engine = create_engine('postgresql+psycopg2://', creator=lambda: pg_conn.conn)
        print("Engine created successfully!")
        for year, df in processed_dataframes.items():
            try:
                print(f"Inserting data for year {year}...")
                df.to_sql(
                    name='workout_journal', # table name
                    con=engine,
                    schema='raw_data',
                    index=False,  # don't include the df's index as a column
                    if_exists='append',  # append, as overwrite would only save the latest year
                    method='multi',  # inserting multiple rows at once to optimize performance
                )
                print(f"Year {year} data inserted successfully!")
            except Exception as e:
                print(f"❌ Error inserting data for year {year}: {e}")
                sys.exit(1)


def process_the_workout_dictionary(final_dictionary: Dict[str, List[List]]) -> Optional[Dict[str, pd.DataFrame]]:
    """
    Processes the final dictionary into pandas DataFrames with necessary adjustments.

    Args:
        final_dictionary (Dict[str, List[List]]): The dictionary containing workout data for each year.

    Returns:
        Optional[Dict[str, pd.DataFrame]]: A dictionary of pandas DataFrames, each corresponding to a year of workout data.
    """
    processed_dataframes = {}
    for year in final_dictionary.keys():
        # creating the dataframe with the desired columns
        print(f"Loading year {year} into a pandas dataframe...")
        workout_df = pd.DataFrame(final_dictionary[year], columns=['month', 'weekday', 'date', 'deload', 'workout_counter', 'set_counter', 'type', 'section', 'exercise', 'weight', 'grip', 'amount', 'rest'])
        
        # creating the weight_unit column, to better aggregate data with the dbt models
        workout_df[["weight", "weight_unit"]] = workout_df["weight"].apply(lambda x: pd.Series(extract_weight_kg(x)))

        # modifying the weight column to have 2 decimal places
        workout_df['weight'] = workout_df['weight'].round(2)

        # setting the desired column order
        desired_column_order = ['month', 'weekday', 'date', 'deload', 'workout_counter', 'set_counter', 'type', 'section', 'exercise', 'weight', 'weight_unit', 'grip', 'amount', 'rest']
        workout_df = workout_df[desired_column_order]

        # applying some changes to the columns, such as setting them to lowercase or modifying date formats, etc
        workout_df['date'] = pd.to_datetime(workout_df['date'], format='%d-%m-%Y').dt.strftime('%Y-%m-%d')
        workout_df['workout_counter'] = workout_df['workout_counter'].str.replace('/365', '', regex=False)
        workout_df['month'] = workout_df['month'].str.lower()
        workout_df['weekday'] = workout_df['weekday'].str.lower()
        workout_df['type'] = workout_df['type'].str.lower()
        workout_df['section'] = workout_df['section'].str.lower()
        workout_df['exercise'] = workout_df['exercise'].str.lower()
        workout_df['weight_unit'] = workout_df['weight_unit'].str.lower()
        workout_df['grip'] = workout_df['grip'].str.lower()

        # adding the dataframe to a dictionary
        processed_dataframes[year] = workout_df
        print(f"Year {year} successfully loaded into a pandas dataframe!")
    
    return processed_dataframes if processed_dataframes else None


if __name__ == '__main__':
    try:
        # schema and table names
        schema="raw_data"
        table_name="workout_journal"

        # workout sheet download link, and names for converting from xlsx to xls
        workout_sheet_xls = "workout_sheet.xls"
        workout_sheet_xlsx = "workout_sheet.xlsx"
        workout_sheet_url = os.getenv('WORKOUT_SHEET_URL')
        
        # make sure the output folder exists before writing to it
        create_output_folder()

        # download the worksheet
        if workout_sheet_url:
            download_workout_sheet(workout_sheet_url=workout_sheet_url,workout_sheet_xlsx=workout_sheet_xlsx)
        else:
            print("Wasn't able to retrieve the workout sheet URL. Please check the environment variables.")
            quit()
        
        # convert the worksheet
        convert_xlsx_to_xls(xlsx_file=workout_sheet_xlsx,xls_file=workout_sheet_xls)
        
        # processing the worksheet
        final_dictionary = read_workbook(workout_sheet=workout_sheet_xls)
        processed_dataframes = process_the_workout_dictionary(final_dictionary=final_dictionary)

        # insert the information into postgres
        insert_dataframes_into_postgres(processed_dataframes=processed_dataframes,schema=schema,table_name=table_name)
    except Exception as e:
        print(f"Something went wrong with the workout_journal script! -> {e}")
        sys.exit(1)
